import openpyxl

path="C:\\Users\\malli\\Downloads\\Book2.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook.active


for r in range(1,8):
    for c in range(1,3):
       sheet.cell(row=r, column=c).value="product Name"

for r in range(2,8):
    for c in range(1,2):
       sheet.cell(row=r, column=c).value="POTATO"

for r in range(3,8):
    for c in range(1,2):
       sheet.cell(row=r, column=c).value="MILK SHAKE"

for r in range(4,8):
    for c in range(1,2):
       sheet.cell(row=r, column=c).value="MAGGIE"

for r in range(5,8):
    for c in range(1,2):
       sheet.cell(row=r, column=c).value="NESCAFE"

for r in range(6,8):
    for c in range(1,2):
       sheet.cell(row=r, column=c).value="OREO Biscuits"

for r in range(7,8):
    for c in range(1,2):
       sheet.cell(row=r, column=c).value="KITKAT "

for r in range(1,2):
    for c in range(2,3):
       sheet.cell(row=r, column=c).value="Quantity"

for r in range(2,3):
    for c in range(2,3):
       sheet.cell(row=r, column=c).value="1"

for r in range(3,4):
    for c in range(2,3):
       sheet.cell(row=r, column=c).value="2"

for r in range(4,5):
    for c in range(2,3):
       sheet.cell(row=r, column=c).value="1"

for r in range(5,6):
    for c in range(2,3):
       sheet.cell(row=r, column=c).value="3"


for r in range(6,7):
    for c in range(2,3):
       sheet.cell(row=r, column=c).value="2"

for r in range(7,8):
    for c in range(2,3):
       sheet.cell(row=r, column=c).value="4"




workbook.save(path)