import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from selenium.webdriver import ActionChains
import unittest
import HtmlTestRunner

path = "C:\\Users\\malli\\Downloads\\Book1.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.get_sheet_by_name("loginpage")
site = sheet['B1'].value
mobileno = sheet['B2'].value
gmail = sheet['B3'].value

driver = webdriver.Chrome(executable_path="C:\\Users\\malli\\Downloads\\chromedriver_win32\\chromedriver.exe")
driver.implicitly_wait(10)
driver.maximize_window()

driver.get(site)

driver.find_element(by=By.NAME, value="otpEmail").send_keys(mobileno)

driver.find_element(by=By.XPATH, value="//login/div/form/div[2]/button[3]").click()
driver.implicitly_wait(10)
print("PASSED")
sheet.cell(2, 3).value = "PASSED"
workbook.save(path)

driver.find_element(by=By.XPATH, value="//span[@class='login-icon login-icon-back-nav']").click()
driver.find_element(by=By.XPATH, value="(//button[@type='button'])[2]").click()
driver.find_element(by=By.ID, value="otpEmail").send_keys(gmail)

driver.find_element(by=By.XPATH, value="(//button[@type='submit'])[1]").click()
sheet.cell(3,3).value="PASSED"
workbook.save(path)

